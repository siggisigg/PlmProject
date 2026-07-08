"""
Electrical BOM ingest (SEE Electrical export).

Usage:
    python -m ingest.elec_bom <path-to-ELECTRICAL_xxx_BOM.xlsx>

Column layout (row 1 = headers):
  Manufacturer, Equipment, Type Description, Goods Group,
  Number of Pieces, Length, Quantity, …, Price, …, Unit price, …

The file has ~17 summary/aggregate rows at the top (no Manufacturer/TypeDesc).
Real component rows have both Manufacturer and Type Description populated.

Part identifier: Equipment (manufacturer catalog number).
All parts get bom_type='electrical'.
BOM is flat — no assembly hierarchy; all bom_lines have parent_line_id = NULL.
A synthetic root part "ELEC-<ref>" is created from the filename stem.

Price/unit_price columns are mostly 0; only non-zero prices are written to
part_revisions to avoid clobbering Vöruskrá data.

Idempotent: if bom_headers row for source_file already exists, BOM lines
are not re-inserted.
"""
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from .db import get_conn, write_sync_log

HEADER_ROW = 1
DATA_START  = 2


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _ref_from_filename(name: str) -> str:
    """Extract the reference code from the filename.
    ELECTRICAL_3155f02_BOM.xlsx → '3155f02'
    Falls back to the full stem if pattern does not match.
    """
    stem = Path(name).stem  # e.g. "ELECTRICAL_3155f02_BOM"
    # Try ELECTRICAL_<ref>_BOM or ELECTRICAL_<ref>
    m = re.match(r'ELECTRICAL[_-](.+?)(?:[_-]BOM)?$', stem, re.IGNORECASE)
    return m.group(1) if m else stem


def run(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    print(f"[elec_bom] Loading {path.name} …")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    col = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}

    source_file = path.name
    ref_code    = _ref_from_filename(source_file)
    root_pn     = f"ELEC-{ref_code}"

    parts_added = parts_updated = bom_lines_added = staged = skipped = 0

    all_rows = list(ws.iter_rows(min_row=DATA_START, values_only=True))
    wb.close()

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            # ── Stage all rows ──────────────────────────────────────────────
            for row_num, row in enumerate(all_rows, start=DATA_START):
                def g(name):
                    idx = col.get(name)
                    return row[idx] if idx is not None and idx < len(row) else None

                cur.execute(
                    """
                    INSERT INTO _staging.electrical_bom_raw
                      (source_file, row_num, manufacturer, equipment, type_description,
                       goods_group, num_pieces, quantity, length_m,
                       price, unit_price_raw, order_number, supplier)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        source_file, row_num,
                        str(g("Manufacturer") or "").strip() or None,
                        str(g("Equipment") or "").strip() or None,
                        str(g("Type Description") or "").strip() or None,
                        str(g("Goods Group") or "").strip() or None,
                        str(g("Number of Pieces") or "").strip() or None,
                        str(g("Quantity") or "").strip() or None,
                        str(g("Length") or "").strip() or None,
                        str(g("Price") or "").strip() or None,
                        str(g("Unit price") or "").strip() or None,
                        str(g("Order number") or "").strip() or None,
                        str(g("Supplier") or "").strip() or None,
                    ),
                )
                staged += 1

            # ── Check idempotency ───────────────────────────────────────────
            cur.execute(
                "SELECT id FROM bom_headers WHERE source_file = %s LIMIT 1",
                (source_file,),
            )
            if cur.fetchone():
                print(f"[elec_bom] {source_file!r} already imported — skipping BOM lines.")
                conn.commit()
                write_sync_log(conn, source_file, 0, 0, len(all_rows), "Already imported.")
                return

            # ── Ensure synthetic root part exists ───────────────────────────
            cur.execute(
                """
                INSERT INTO parts (plm_part_number, description, is_assembly,
                                   is_master_assembly, bom_type)
                VALUES (%s, %s, true, true, 'electrical')
                ON CONFLICT (plm_part_number) DO UPDATE
                  SET description = EXCLUDED.description
                RETURNING id
                """,
                (root_pn, f"Electrical assembly — {ref_code}"),
            )
            root_part_id = cur.fetchone()[0]

            # ── Create bom_headers row ──────────────────────────────────────
            cur.execute(
                """
                INSERT INTO bom_headers (root_part_id, source_file, bom_type)
                VALUES (%s, %s, 'electrical')
                RETURNING id
                """,
                (root_part_id, source_file),
            )
            header_id = cur.fetchone()[0]

            # ── Upsert component parts and insert flat bom_lines ───────────
            sort_order = 0
            for row in all_rows:
                def g(name):
                    idx = col.get(name)
                    return row[idx] if idx is not None and idx < len(row) else None

                manufacturer = str(g("Manufacturer") or "").strip() or None
                equipment    = str(g("Equipment") or "").strip() or None
                type_desc    = str(g("Type Description") or "").strip() or None

                # Skip summary/aggregate rows that have no real component data
                if not manufacturer and not type_desc:
                    skipped += 1
                    continue
                if not equipment:
                    skipped += 1
                    continue

                goods_group = str(g("Goods Group") or "").strip() or None
                num_pieces  = _to_decimal(g("Number of Pieces"))
                quantity    = _to_decimal(g("Quantity"))
                price_raw   = _to_decimal(g("Price"))
                unit_p      = _to_decimal(g("Unit price"))

                # Use quantity if num_pieces is None, default to 1
                qty = num_pieces or quantity or Decimal("1")

                # Part identifier: Equipment (catalog/order number)
                part_pn = equipment

                cur.execute(
                    """
                    INSERT INTO parts
                      (plm_part_number, description, manufacturer, bom_type,
                       is_assembly, erp_category)
                    VALUES (%s, %s, %s, 'electrical', false, %s)
                    ON CONFLICT (plm_part_number) DO UPDATE
                      SET description = EXCLUDED.description,
                          manufacturer = EXCLUDED.manufacturer,
                          erp_category = EXCLUDED.erp_category
                    RETURNING id, (xmax = 0) AS inserted
                    """,
                    (part_pn, type_desc, manufacturer, goods_group),
                )
                part_id, was_inserted = cur.fetchone()
                if was_inserted:
                    parts_added += 1
                else:
                    parts_updated += 1

                # Only write a price revision if we have a non-zero unit price
                effective_price = unit_p if (unit_p and unit_p > 0) else (
                    price_raw if (price_raw and price_raw > 0) else None
                )
                if effective_price:
                    cur.execute(
                        """
                        SELECT id FROM part_revisions
                        WHERE part_id = %s AND valid_to IS NULL LIMIT 1
                        """,
                        (part_id,),
                    )
                    if not cur.fetchone():
                        cur.execute(
                            """
                            INSERT INTO part_revisions (part_id, unit_price, currency)
                            VALUES (%s, %s, 'ISK')
                            """,
                            (part_id, effective_price),
                        )

                # Flat BOM line — all are direct children of the root (parent_line_id = NULL)
                cur.execute(
                    """
                    INSERT INTO bom_lines
                      (bom_header_id, parent_line_id, part_id, quantity, sort_order)
                    VALUES (%s, NULL, %s, %s, %s)
                    """,
                    (header_id, part_id, qty, sort_order),
                )
                bom_lines_added += 1
                sort_order += 1

        conn.commit()
        write_sync_log(
            conn, source_file,
            parts_added, parts_updated, skipped,
            (f"Elec BOM: {parts_added} new parts, {parts_updated} updated, "
             f"{skipped} summary rows skipped, {bom_lines_added} BOM lines"),
        )

    finally:
        conn.close()

    print(f"[elec_bom] Done — {parts_added} parts added, {parts_updated} updated, "
          f"{skipped} rows skipped, {bom_lines_added} BOM lines")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m ingest.elec_bom <path-to-xlsx>")
        sys.exit(1)
    run(sys.argv[1])
