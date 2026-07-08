"""
Mechanical BOM ingest (SolidWorks PDM export).

Usage:
    python -m ingest.mech_bom <path-to-BOM.xlsx>

Column layout (row 1 = headers, row 2+ = data):
  Level, Document Preview, Part Number, Description, Quantity, Weight,
  Material, Description - Blank, Part number - Blank, Length,
  Sheet metal thickness, Production Type, Manufacturer, Code, Revision, Comment

Level column quirks:
  - Top-level row: integer 1
  - Two-segment levels (1.1, 1.2 …): Python float
  - Deeper levels (1.3.1 …): already a string
  All are normalized to dotted decimal strings ("1", "1.1", "1.3.1").

Cut-list items: Part Number contains "<N>" (e.g. "Threaded rod<2>").
  Staged but NOT written to parts/bom_lines.

Assembly detection: Production Type contains "(SS)".
Master Assembly: the single root row at level "1".

Idempotent: if a bom_headers row for (source_file, root_part_number, revision)
already exists, BOM lines are not re-inserted.
"""
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from .db import get_conn, write_sync_log

HEADER_ROW = 1
DATA_START  = 2

_CUT_LIST_RE = re.compile(r'<\d+>')
_PROD_TYPE_RE = re.compile(r'\(([A-Z]+)\)')


def _normalize_level(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        # Avoid floating-point artifacts: 1.1000000000000001 → "1.1"
        return f"{val:.10g}"
    return str(val).strip()


def _parent_level(level: str) -> str | None:
    parts = level.split(".")
    return ".".join(parts[:-1]) if len(parts) > 1 else None


def _parse_prod_type(val) -> str | None:
    if not val:
        return None
    m = _PROD_TYPE_RE.search(str(val))
    return m.group(1) if m else str(val).strip()


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def run(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    print(f"[mech_bom] Loading {path.name} …")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Build column index map from header row
    header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    col = {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}

    source_file = path.name
    parts_added = parts_updated = bom_lines_added = staged = skipped_cut = 0

    # --- Read all data rows into memory first (avoid nested cursor issues)
    all_rows = list(ws.iter_rows(min_row=DATA_START, values_only=True))
    wb.close()

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            # ── Stage all rows ──────────────────────────────────────────────
            for row_num, row in enumerate(all_rows, start=DATA_START):
                level_raw   = row[col.get("Level", 0)] if "Level" in col else None
                part_num    = row[col.get("Part Number", 2)] if "Part Number" in col else None
                description = row[col.get("Description", 3)] if "Description" in col else None
                quantity    = row[col.get("Quantity", 4)] if "Quantity" in col else None
                weight      = row[col.get("Weight", 5)] if "Weight" in col else None
                material    = row[col.get("Material", 6)] if "Material" in col else None
                mat_shape   = row[col.get("Description - Blank", 7)] if "Description - Blank" in col else None
                mat_spec    = row[col.get("Part number - Blank", 8)] if "Part number - Blank" in col else None
                length_mm   = row[col.get("Length", 9)] if "Length" in col else None
                thickness   = row[col.get("Sheet metal thickness", 10)] if "Sheet metal thickness" in col else None
                prod_type   = row[col.get("Production Type", 11)] if "Production Type" in col else None
                manufacturer = row[col.get("Manufacturer", 12)] if "Manufacturer" in col else None
                revision    = row[col.get("Revision", 14)] if "Revision" in col else None
                comment     = row[col.get("Comment", 15)] if "Comment" in col else None

                part_num_str = str(part_num).strip() if part_num else None
                is_cut = bool(part_num_str and _CUT_LIST_RE.search(part_num_str))

                cur.execute(
                    """
                    INSERT INTO _staging.parts_raw
                      (source_file, row_num, plm_part_number, description, material,
                       material_shape, material_spec, length_mm, thickness_mm,
                       production_type, manufacturer, revision, comment, is_cut_list_item)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        source_file, row_num,
                        part_num_str if not is_cut else None,
                        str(description).strip() if description else None,
                        str(material).strip() if material else None,
                        str(mat_shape).strip() if mat_shape else None,
                        str(mat_spec).strip() if mat_spec else None,
                        str(length_mm) if length_mm is not None else None,
                        str(thickness) if thickness is not None else None,
                        str(prod_type).strip() if prod_type else None,
                        str(manufacturer).strip() if manufacturer else None,
                        str(revision).strip() if revision else None,
                        str(comment).strip() if comment else None,
                        is_cut,
                    ),
                )
                staged += 1

            # ── Identify root assembly (level "1") ──────────────────────────
            root_row = None
            for row in all_rows:
                level_val = row[col.get("Level", 0)] if "Level" in col else None
                lvl = _normalize_level(level_val)
                if lvl == "1":
                    root_row = row
                    break

            if root_row is None:
                raise ValueError("Could not find root row at Level '1' in BOM.")

            root_part_num = str(root_row[col.get("Part Number", 2)] or "").strip()
            root_revision = str(root_row[col.get("Revision", 14)] or "").strip() or None

            if not root_part_num:
                raise ValueError("Root row (Level 1) has no Part Number.")

            # ── Check if this exact BOM has already been imported ───────────
            cur.execute(
                """
                SELECT id FROM bom_headers
                WHERE source_file = %s AND revision IS NOT DISTINCT FROM %s
                LIMIT 1
                """,
                (source_file, root_revision),
            )
            if cur.fetchone():
                print(f"[mech_bom] BOM {source_file!r} revision {root_revision!r} "
                      f"already imported — skipping BOM lines rebuild.")
                conn.commit()
                write_sync_log(conn, source_file, 0, 0, len(all_rows),
                               "Already imported — no changes.")
                return

            # ── Upsert all parts ────────────────────────────────────────────
            part_id_map: dict[str, int] = {}

            for row in all_rows:
                level_val   = row[col.get("Level", 0)] if "Level" in col else None
                part_num    = row[col.get("Part Number", 2)] if "Part Number" in col else None
                description = row[col.get("Description", 3)] if "Description" in col else None
                material    = row[col.get("Material", 6)] if "Material" in col else None
                manufacturer = row[col.get("Manufacturer", 12)] if "Manufacturer" in col else None
                prod_type   = row[col.get("Production Type", 11)] if "Production Type" in col else None
                revision    = row[col.get("Revision", 14)] if "Revision" in col else None

                part_num_str = str(part_num).strip() if part_num else None
                if not part_num_str:
                    continue
                if _CUT_LIST_RE.search(part_num_str):
                    skipped_cut += 1
                    continue

                lvl = _normalize_level(level_val)
                prod_code = _parse_prod_type(prod_type)
                is_assembly = prod_code == "SS"
                is_master   = lvl == "1"
                desc_clean  = str(description).strip() if description else None

                cur.execute(
                    """
                    INSERT INTO parts
                      (plm_part_number, description, material, production_type,
                       manufacturer, is_assembly, is_master_assembly, bom_type)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,'mechanical')
                    ON CONFLICT (plm_part_number) DO UPDATE
                      SET description        = EXCLUDED.description,
                          material           = EXCLUDED.material,
                          production_type    = EXCLUDED.production_type,
                          manufacturer       = EXCLUDED.manufacturer,
                          is_assembly        = EXCLUDED.is_assembly,
                          is_master_assembly = EXCLUDED.is_master_assembly
                    RETURNING id, (xmax = 0) AS inserted
                    """,
                    (
                        part_num_str, desc_clean,
                        str(material).strip() if material else None,
                        prod_code,
                        str(manufacturer).strip() if manufacturer else None,
                        is_assembly, is_master,
                    ),
                )
                part_id, was_inserted = cur.fetchone()
                part_id_map[part_num_str] = part_id
                if was_inserted:
                    parts_added += 1
                else:
                    parts_updated += 1

                # Insert part_revision if no current row exists (cost comes from Vöruskrá)
                if not is_assembly and revision:
                    cur.execute(
                        """
                        SELECT id FROM part_revisions
                        WHERE part_id = %s AND valid_to IS NULL LIMIT 1
                        """,
                        (part_id,),
                    )
                    if not cur.fetchone():
                        cur.execute(
                            """
                            INSERT INTO part_revisions (part_id, revision, currency)
                            VALUES (%s, %s, 'ISK')
                            """,
                            (part_id, str(revision).strip()),
                        )

            # ── Create bom_headers row ──────────────────────────────────────
            root_part_id = part_id_map.get(root_part_num)
            if root_part_id is None:
                raise ValueError(f"Root part {root_part_num!r} was not inserted — check data.")

            cur.execute(
                """
                INSERT INTO bom_headers (root_part_id, source_file, bom_type, revision)
                VALUES (%s, %s, 'mechanical', %s)
                RETURNING id
                """,
                (root_part_id, source_file, root_revision),
            )
            header_id = cur.fetchone()[0]

            # ── Insert bom_lines (depth-stack approach) ────────────────────
            # Stack stores (depth, line_id) of current ancestor chain.
            # Robust against float normalization collisions (e.g. 8.1 vs 8.10).
            # Works because the BOM export is depth-first ordered.
            depth_stack: list[tuple[int, int]] = []

            for sort_order, row in enumerate(all_rows):
                level_val = row[col.get("Level", 0)] if "Level" in col else None
                lvl = _normalize_level(level_val)
                if lvl is None:
                    continue

                part_num = row[col.get("Part Number", 2)] if "Part Number" in col else None
                part_num_str = str(part_num).strip() if part_num else None
                if not part_num_str or _CUT_LIST_RE.search(part_num_str):
                    continue

                part_id = part_id_map.get(part_num_str)
                if part_id is None:
                    continue

                quantity = row[col.get("Quantity", 4)] if "Quantity" in col else None
                qty = _to_decimal(quantity) or Decimal("1")

                depth = len(lvl.split("."))

                # Pop stack until the top is at depth - 1 (our direct parent)
                while depth_stack and depth_stack[-1][0] >= depth:
                    depth_stack.pop()

                parent_line_id = depth_stack[-1][1] if depth_stack else None

                cur.execute(
                    """
                    INSERT INTO bom_lines
                      (bom_header_id, parent_line_id, part_id, quantity, level_path, sort_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (header_id, parent_line_id, part_id, qty, lvl, sort_order),
                )
                line_id = cur.fetchone()[0]
                depth_stack.append((depth, line_id))
                bom_lines_added += 1

        conn.commit()
        write_sync_log(
            conn, source_file,
            parts_added, parts_updated, skipped_cut,
            (f"Mech BOM: {parts_added} new parts, {parts_updated} updated, "
             f"{skipped_cut} cut-list skipped, {bom_lines_added} BOM lines inserted"),
        )

    finally:
        conn.close()

    print(f"[mech_bom] Done — {parts_added} parts added, {parts_updated} updated, "
          f"{skipped_cut} cut-list items skipped, {bom_lines_added} BOM lines")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m ingest.mech_bom <path-to-xlsx>")
        sys.exit(1)
    run(sys.argv[1])
