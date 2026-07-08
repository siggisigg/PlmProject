"""
Vöruskrá (ERP price/item register) ingest.

Usage:
    python -m ingest.voruskra <path-to-Vöruskrá.xlsx>

The file has a group-header row 1 ("Almennt") and column headers in row 2.
Data rows start at row 3.

For each row:
  - Upserts parts (plm_part_number, description, unit_of_measure, erp_category)
  - Upserts part_revisions: if current cost+price unchanged → skip;
    if changed → close old revision, open new one.

Safe to re-run on the same file (idempotent).
"""
import sys
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

_DB_SCALE = Decimal("0.0001")  # NUMERIC(14,4) — match DB rounding when comparing


def _round4(d: Decimal | None) -> Decimal:
    if d is None:
        return Decimal("0")
    return d.quantize(_DB_SCALE, rounding=ROUND_HALF_UP)
from pathlib import Path

import openpyxl

from .db import get_conn, write_sync_log


HEADER_ROW = 2   # 1-indexed; row 1 is group header "Almennt"
DATA_START  = 3


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    return None


def load_headers(ws) -> dict[str, int]:
    """Return {column_name: 0-based-index} from the header row."""
    row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return {str(v).strip(): i for i, v in enumerate(row) if v is not None}


def run(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    print(f"[voruskra] Loading {path.name} …")

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    col = load_headers(ws)

    # Required columns
    for name in ("Vörunúmer", "Kostnaðarverð", "Verð 1"):
        if name not in col:
            raise ValueError(f"Expected column '{name}' not found in {path.name}. "
                             f"Found: {list(col)}")

    source_file = path.name
    added = updated = skipped = 0

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for row_values in ws.iter_rows(min_row=DATA_START, values_only=True):
                vorunumer = row_values[col["Vörunúmer"]]
                if not vorunumer:
                    continue
                vorunumer = str(vorunumer).strip()
                if not vorunumer:
                    continue

                # Pull columns, tolerating missing ones
                def get(name):
                    idx = col.get(name)
                    return row_values[idx] if idx is not None else None

                desc1 = str(get("Vörulýsing") or "").strip()
                desc2 = str(get("Vörulýsing 2") or "").strip()
                description = (desc1 + (" — " + desc2 if desc2 else "")).strip(" —") or None

                unit_cost  = _to_decimal(get("Kostnaðarverð"))
                unit_price = _to_decimal(get("Verð 1"))
                uom        = str(get("Grunneining") or "").strip() or None
                category   = str(get("Vöruflokkur") or "").strip() or None
                modified   = _to_date(get("Breytt dags."))

                # Stage raw row
                cur.execute(
                    """
                    INSERT INTO _staging.voruskra_raw
                      (source_file, vorunumer, vorulysingur, vorulysingur_2,
                       kostnadarverd, verd_1, grunneining, voruflokk, breytt_dags)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (source_file, vorunumer, desc1 or None, desc2 or None,
                     unit_cost, unit_price, uom, category, modified),
                )

                # Upsert parts
                cur.execute(
                    """
                    INSERT INTO parts (plm_part_number, description, unit_of_measure, erp_category)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (plm_part_number) DO UPDATE
                      SET description     = EXCLUDED.description,
                          unit_of_measure = EXCLUDED.unit_of_measure,
                          erp_category    = EXCLUDED.erp_category
                    RETURNING id, (xmax = 0) AS inserted
                    """,
                    (vorunumer, description, uom, category),
                )
                part_id, inserted = cur.fetchone()

                # Upsert part_revisions: compare with current (valid_to IS NULL)
                cur.execute(
                    """
                    SELECT id, unit_cost, unit_price
                    FROM part_revisions
                    WHERE part_id = %s AND valid_to IS NULL
                    ORDER BY valid_from DESC
                    LIMIT 1
                    """,
                    (part_id,),
                )
                current = cur.fetchone()

                if current is None:
                    # No revision yet — insert first one
                    cur.execute(
                        """
                        INSERT INTO part_revisions (part_id, unit_cost, unit_price, currency)
                        VALUES (%s, %s, %s, 'ISK')
                        """,
                        (part_id, unit_cost, unit_price),
                    )
                    added += 1
                else:
                    rev_id, existing_cost, existing_price = current
                    same = (
                        _round4(_to_decimal(existing_cost)) == _round4(unit_cost) and
                        _round4(_to_decimal(existing_price)) == _round4(unit_price)
                    )
                    if same:
                        skipped += 1
                    else:
                        # Close old revision, open new one
                        cur.execute(
                            "UPDATE part_revisions SET valid_to = now() WHERE id = %s",
                            (rev_id,),
                        )
                        cur.execute(
                            """
                            INSERT INTO part_revisions (part_id, unit_cost, unit_price, currency)
                            VALUES (%s, %s, %s, 'ISK')
                            """,
                            (part_id, unit_cost, unit_price),
                        )
                        updated += 1

        conn.commit()
        write_sync_log(conn, source_file, added, updated, skipped,
                       f"Vöruskrá ingest — {added} new, {updated} price changes, {skipped} unchanged")

    finally:
        wb.close()
        conn.close()

    print(f"[voruskra] Done — {added} added, {updated} updated, {skipped} unchanged")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m ingest.voruskra <path-to-xlsx>")
        sys.exit(1)
    run(sys.argv[1])
